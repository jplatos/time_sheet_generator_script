from itertools import product
from openpyxl import worksheet
from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import os
from datetime import datetime
from person import Person, MonthRecord
from loading import load_data, load_datas, create_file_names, month_hours, month_state_holidays_hours, month_names, report_input_file, report_attendance_file, report_months, month_days, state_holidays, day_names_short


def fill_row(ws, day):
	ws.cell(row = 12+day, column=1).fill = PatternFill("solid", fgColor="DDDDDD")
	ws.cell(row = 12+day, column=2).fill = PatternFill("solid", fgColor="DDDDDD")
	ws.cell(row = 12+day, column=3).fill = PatternFill("solid", fgColor="DDDDDD")
	ws.cell(row = 12+day, column=4).fill = PatternFill("solid", fgColor="DDDDDD")
	ws.cell(row = 12+day, column=15).fill = PatternFill("solid", fgColor="DDDDDD")

def export_timesheets(persons, file_name):

	for person_idx, person in enumerate(persons):
		
		print(f"{person_idx:5d} Year={person.report_year:4d} Month={person.report_month:02d} Name={person.nick}")
		rec = person.records[0]
		
		# read a output form
		res_wb = load_workbook(file_name)
		ws = res_wb.active
		# name
		ws['D7'].value = person.nick
		ws['K7'].value = person.key_activity
		ws['D8'].value = person.position
		ws['K8'].value = rec.contract_type

		
		if rec.is_contract:
			if (rec.projectTotal!=0):
				# amount of work in the project
				
				ws['D9'].value = rec.projectAmount_str
		else:
			ws['D9'].value = rec.projectTotal
			
		ws['K9'].value = f'{person.report_month:02}_{person.report_year}'
		
		
		# mark weekends
		for day in range(1,month_days[person.report_month]+1):
			day_of_week = datetime(person.report_year, person.report_month, day).weekday()
			ws.cell(column=1, row=12+day).value = "{}. {}".format(day, day_names_short[day_of_week])
			if day_of_week>4:
				fill_row(ws, day)
		# mark state holidays
		for day in state_holidays[person.report_month]:
			day_of_week = datetime(person.report_year, person.report_month, day).weekday()
			if day_of_week<5:
				ws.cell(row=12+day, column=2).value = 'Státní svátek'
				fill_row(ws, day)
		
		for day in range(month_days[person.report_month], 31):
			ws.row_dimensions[13+day].hidden = True

		rec = person.records[0]
		for day,value in rec.attendance.items():
			if value=='d':
				ws.cell(row=12+day, column=2).value = 'Půl dne dovolené'
			if value=='D':
				fill_row(ws, day)
				ws.cell(row=12+day, column=2).value = 'Dovolená'
			if value=='C':
				fill_row(ws, day)
				ws.cell(row=12+day, column=2).value = 'Ostatní překážky v práci / indispoziční volno'
			if value=='*':
				fill_row(ws, day)
				ws.cell(row=12+day, column=2).value = 'Ostatní překážky v práci / indispoziční volno'			
			if value=='N':
				fill_row(ws, day)
				ws.cell(row=12+day, column=2).value = 'Pracovní neschopnost'
			if value=='P' or value=='O':
				fill_row(ws, day)
				ws.cell(row=12+day, column=2).value = 'Překážka v práci'
			if value=='p' or value=='o':
				fill_row(ws, day)
				ws.cell(row=12+day, column=2).value = 'Překážka v práci/část'
			

		
		ws['E49'].value = rec.holidays_str()
		ws['E50'].value = rec.holidays_days()*8*rec.projectAmount
		ws['E51'].value = rec.holidays_days()*8*rec.projectAmount

		ws['E54'].value = rec.illnesses_str()
		ws['E55'].value = rec.illnesses_days()*8*rec.projectAmount
		ws['E56'].value = rec.illnesses_days()*8*rec.projectAmount

		if rec.is_contract and len(state_holidays[person.report_month]):			
			days = [x for x in state_holidays[person.report_month] if datetime(person.report_year, person.report_month, x).weekday()<5]
			ws['M49'].value = ', '.join(str(x) for x in sorted(days))
			ws['M50'].value = len(days)*8*rec.projectAmount
			ws['M51'].value = len(days)*8*rec.projectAmount
		else:
			ws['M50'].value = 0
			ws['M51'].value = 0

		ws['M54'].value = rec.obstacles_str()
		ws['M55'].value = rec.obstacles_days()*8*rec.projectAmount
		ws['M56'].value = rec.obstacles_days()*8*rec.projectAmount

		# celkovy pocet hodin
		if (rec.is_contract):
			total_hours = month_hours[person.report_month-1] * rec.projectAmount
			ws['I58'].value = ws['I59'].value = total_hours
		# pocet hodin na DPP/DPC
		else:
			ws['I58'].value = ws['I59'].value = rec.projectTotal

		ws['D67'].value = person.nick
		ws['D68'].value = person.approver
		ws['I68'].value = person.approver_position
		
		# store to the file
		base_folder = "output"
		ka_folder = person.key_activity
		spp_folder = person.spp
		out_file_name = person.file_name
		if ka_folder is None:
			ka_folder = "KA"
		folder = base_folder+'/'+ka_folder+'/'+spp_folder
		if not os.path.exists(folder):
			os.makedirs(folder)
		
		res_wb.save(folder+"/"+out_file_name)	

		# if person_idx==0: break



persons = load_datas()
	
print(len(persons), "Persons loaded")

file_name = "input/TEMPLATE"

export_timesheets(persons, file_name)

