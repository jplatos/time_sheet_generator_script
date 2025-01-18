from itertools import product
import types
import openpyxl
from openpyxl import worksheet
from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import os
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from person import Person, MonthRecord

day_names = ['pondělí', 'úterý', 'středa', 'čtvrtek', 'pátek','sobota','neděle']
day_names_short = ['Po', 'Út', 'Stř', 'Čt', 'Pá','So','Ne']
month_names = ['leden', 'únor', 'březen', 'duben', 'květen', 'červen', 'červenec', 'srpen', 'září', 'říjen', 'listopad', 'prosinec']
month_hours = [184, 168, 168, 176, 184, 160, 184, 176, 168, 184, 168, 176]
month_state_holidays_hours = [8, 0, 8, 8, 16, 0, 8, 0, 0, 8, 0, 24]

month_days = {1:31, 2:29, 3:31, 4:30, 5:31, 6:30, 7:31, 8:31, 9:30, 10:31, 11:30, 12:31}
state_holidays = {	1:set([1]), 2:set(), 3:set([29]), 4:set([1]), 5:set([1,8]), 6:set(), 7:set([5,6]), 8:set(), 9:set([28]), 10:set([28]), 11:set([17]), 12:set([24, 25, 26]) }


report_months = [(2024, 9), (2024, 10), (2024, 11), (2024, 12)]

report_input_file = f"input/XXX.xlsx"
report_attendance_file = f"input/XXX.xlsx"

def is_number(s):
	try:
		float(s)
		return True
	except ValueError:
		return False


def load_sap_report(file_name, sheet_name, year, month):
	data_wb = load_workbook(file_name, data_only=True)
	data_ws = data_wb[sheet_name]
	actual_row = 2
	sap_no_col = column_index_from_string('A')
	day_col_start = column_index_from_string('C')-1

	work_days = [x for x in range(1, month_days[month]+1) if datetime(year, month, x).weekday()<5 and not x in state_holidays[month] ]

	attendance = {}
	while (data_ws.cell(column=sap_no_col, row = actual_row).value!=None):
		sap_no = data_ws.cell(row=actual_row, column = sap_no_col).value
		att_days = {}
		for day in work_days:
			val = data_ws.cell(row=actual_row, column = day_col_start+day).value
			if val:
				att_days[day] = val.strip()
		attendance[sap_no] = att_days

		actual_row += 1
	return attendance
	

def load_data(file_name, sheet_name, sap_report, sap_sheet, report_year, report_month):
	
	attendance = load_sap_report(sap_report,sap_sheet, report_year, report_month)
	persons = []
	data_wb = load_workbook(file_name, data_only=True)
	data_ws = data_wb[sheet_name]
	
	actual_row = 2
	person_nick_column = column_index_from_string('A')
	person_lastname_column = column_index_from_string('B')
	person_firstname_column = column_index_from_string('C')
	person_email_column = column_index_from_string('D')
	person_sap_number_column = column_index_from_string('E')
	person_key_activity_column = column_index_from_string('F')	
	person_spp_column = column_index_from_string('G')
	person_position_column = column_index_from_string('H')
	person_contract_column = column_index_from_string('I')
	person_workload_column = column_index_from_string('K')

	record_work_done_column = column_index_from_string('L')
	obstacle_column = column_index_from_string('P')

	aprover_column = column_index_from_string('U')
	aprover_position_column = column_index_from_string('V')
	
	while (data_ws['A'+str(actual_row)].value!=None):
		person = Person()
		person.report_year = report_year
		person.report_month = report_month
		person.sap_number = str(data_ws.cell(row = actual_row, column = person_sap_number_column).value)
		if person.sap_number=='None': person.sap_number = None
		person.nick = data_ws.cell(row = actual_row, column = person_nick_column).value
		person.lastname = data_ws.cell(row = actual_row, column = person_lastname_column).value
		person.firstname = data_ws.cell(row = actual_row, column = person_firstname_column).value
		person.spp = data_ws.cell(row = actual_row, column = person_spp_column).value
		person.key_activity = data_ws.cell(row = actual_row, column = person_key_activity_column).value
		person.position = data_ws.cell(row = actual_row, column = person_position_column).value
		person.email = data_ws.cell(row = actual_row, column = person_email_column).value
		person.approver = data_ws.cell(row = actual_row, column = aprover_column).value
		person.approver_position = data_ws.cell(row = actual_row, column = aprover_position_column).value

		contract_type = data_ws.cell(row=actual_row, column=person_contract_column).value

		# reading month		
		rec = MonthRecord()
		if person.sap_number is not None:
			sap = int(str(person.sap_number.lstrip('0')))
			if sap in attendance:
				rec.attendance = attendance[sap]
			elif str(sap) in attendance:
				rec.attendance = attendance[str(sap)]
			else:
				rec.attendance = {}
		else:
			rec.attendance = {}
		rec.contract_type = contract_type

		rec.projectAmount = data_ws.cell(row=actual_row, column=person_workload_column).value

		if rec.projectAmount is None:
			rec.work_holyday_hours = 0
			rec.project_holyday_hours = 0				
			rec.project_sickness_hours = 0
			rec.work_sickness_days = 0
			rec.project_obstacles_hours = 0
			rec.work_obstacles_hours = 0
			rec.project_state_holyday_hours = 0
			rec.work_state_holydays_hours = 0
			person.records.append(rec)
			continue

		if not isinstance(rec.projectAmount, float) and not isinstance(rec.projectAmount, int):
			rec.projectAmount = rec.projectAmount.strip().replace(',', '.')
			rec.projectAmount = float(rec.projectAmount)

		if (rec.is_contract):
			# that month does not work
			if rec.projectAmount is None or rec.projectAmount==0:
				rec.work_holyday_hours = 0
				rec.project_holyday_hours = 0
				
				rec.project_sickness_hours = 0
				rec.work_sickness_days = 0
				
				rec.project_obstacles_hours = 0
				rec.work_obstacles_hours = 0

				rec.project_state_holyday_hours = 0
				rec.work_state_holydays_hours = 0
			else:
				rec.projectTotal = month_hours[report_month-1] * rec.projectAmount
		else:
			value = data_ws.cell(row=actual_row, column=record_work_done_column).value
			# non contract - compute by hours
			rec.projectTotal = rec.projectAmount = float(value)	
		if (rec.is_contract):
			rec.obstacles_hours = float(data_ws.cell(row=actual_row, column=obstacle_column).value)
		
		person.records.append(rec)
		
		persons.append(person)

		actual_row += 1

	create_file_names(persons, report_year, report_month)
	return persons

def create_file_names(persons, year, month):
	used_names = set()
	for person in persons:
		count = 1
		fn = "Výkaz práce_{}_{}_{} {}-{:02d}.xlsx".format(person.nick,person.key_activity, person.spp, year, month)
		while fn in used_names:
			fn = "Výkaz práce_{}_{}_{}_{} {}-{:02d}.xlsx".format(person.nick,person.key_activity, person.spp, count, year, month)
			count += 1
		person.file_name = fn
		used_names.add(fn)

def load_datas():
	result = []
	for report_year, report_month in report_months:
		report_input_sheet = f'{report_month:02d}_{report_year}'
		report_attendance_sheet = f'DOCHÁZKA_{report_month:02d}_{report_year}'
		persons = load_data(report_input_file, report_input_sheet, report_attendance_file, report_attendance_sheet, report_year, report_month)
		result.extend(persons)
	return result

if __name__ == '__main__':
	persons = load_datas()
	print(len(persons), "Persons loaded")
	
